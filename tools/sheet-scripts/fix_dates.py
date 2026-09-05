# -*- coding: utf-8 -*-
# Normalise dates to dd/mm/yyyy in ONLY the تصريح and Mail Day columns.
# Everything else — other columns, colours, text values, blanks, merges — is
# left exactly as it is. A cell that is not a parseable date is never touched.
#
# Usage: python fix_dates.py "<path to the downloaded .xlsx>"
import sys, io, re, datetime, json
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else None
if not SRC:
    print("give the .xlsx path as an argument"); sys.exit(1)
OUT = re.sub(r"\.xlsx$", "", SRC) + " - DATES.xlsx"

DATE_RE = re.compile(r'^\s*(\d{1,2})\s*([-/])\s*(\d{1,2})\s*[-/]\s*(\d{4})')

def to_latin(s):
    return (s.replace('٠','0').replace('١','1').replace('٢','2').replace('٣','3')
             .replace('٤','4').replace('٥','5').replace('٦','6').replace('٧','7')
             .replace('٨','8').replace('٩','9'))

def parse_date(v):
    """Return (date, trailing_note) or (None, _). Day-first; month-first only
    when day-first is impossible — matches the app exactly."""
    if v is None:
        return None, ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.date() if isinstance(v, datetime.datetime) else v), ""
    s = to_latin(str(v).strip())
    m = DATE_RE.match(s)
    if not m:
        return None, s
    a, b, y = int(m.group(1)), int(m.group(3)), int(m.group(4))
    if b > 12 and a <= 12:
        month, day = a, b
    else:
        day, month = a, b
    if month < 1 or month > 12 or day < 1 or day > 31:
        return None, s
    try:
        d = datetime.date(y, month, day)
    except ValueError:
        return None, s
    if d.month != month or d.day != day:
        return None, s
    return d, s[m.end():].strip(" -/(),.:;؛")

def fmt(d):
    return "%02d/%02d/%04d" % (d.day, d.month, d.year)

# name col (to skip separator rows), then the date columns to normalise
SHEETS = [
    ("أفراد",  3, [9, 10]),   # تصريح=I, Mali Day=J
    ("مركبات", 3, [6, 7]),    # تصريح=F, Mail Day=G
]

wb = openpyxl.load_workbook(SRC)
report = {}

for keyword, ncol, datecols in SHEETS:
    ws = None
    for sh in wb.worksheets:
        if keyword in sh.title:
            ws = sh; break
    if ws is None:
        report[keyword] = "SHEET NOT FOUND"; continue

    changed = 0; skipped_non_date = 0; separators = 0
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=ncol).value
        if name is None or str(name).strip() == "":
            separators += 1
            continue
        for c in datecols:
            cell = ws.cell(row=r, column=c)
            d, _note = parse_date(cell.value)
            if d:
                new = fmt(d)
                if str(cell.value) != new:
                    cell.value = new
                    cell.number_format = "@"
                    changed += 1
            elif cell.value not in (None, ""):
                skipped_non_date += 1
    report[ws.title] = {"dates_normalised": changed,
                        "non_date_cells_left_alone": skipped_non_date,
                        "separator_rows_skipped": separators}

wb.save(OUT)
print(json.dumps(report, ensure_ascii=False, indent=1))
print("SAVED:", OUT)
