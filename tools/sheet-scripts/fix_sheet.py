# -*- coding: utf-8 -*-
# Normalise dates AND colour the permit column, on the downloaded workbook.
#
# Touches only:  تصريح, Mail Day (dates) ; تصريح colour ; whole row colour for منع.
#   - Every date in تصريح + Mail Day  -> dd/mm/yyyy text (day-first)
#   - Mail Day cells                  -> white (no yellow)
#   - تصريح contains "منع"            -> red cell + ENTIRE ROW red
#   - تصريح == "في انتظار الرد"       -> light grey cell
#   - تصريح is a date, in the future  -> green cell  (سارى)
#   - تصريح is a date, in the past    -> yellow cell (منتهى)
#   - anything else (غير محدد موقف …) -> white
#   - separator rows (blank name)     -> skipped entirely (green kept)
#
# Usage: python fix_sheet.py "<path to the downloaded .xlsx>"
import sys, re, datetime, json
import openpyxl
from openpyxl.styles import PatternFill

SRC = sys.argv[1] if len(sys.argv) > 1 else None
if not SRC:
    print("give the .xlsx path as an argument"); sys.exit(1)
OUT = re.sub(r"\.xlsx$", "", SRC) + " - FIXED.xlsx"

TODAY = datetime.date.today()

RED    = PatternFill("solid", fgColor="FFFF6B6B")
GREEN  = PatternFill("solid", fgColor="FFA9D08E")
YELLOW = PatternFill("solid", fgColor="FFFFD966")
GREY   = PatternFill("solid", fgColor="FFD9D9D9")
WHITE  = PatternFill("solid", fgColor="FFFFFFFF")

DATE_RE = re.compile(r'^\s*(\d{1,2})\s*([-/])\s*(\d{1,2})\s*[-/]\s*(\d{4})')

def to_latin(s):
    out = []
    for ch in s:
        o = ord(ch)
        if 0x0660 <= o <= 0x0669: out.append(chr(o - 0x0660 + 48))
        elif 0x06F0 <= o <= 0x06F9: out.append(chr(o - 0x06F0 + 48))
        else: out.append(ch)
    return "".join(out)

def parse_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.date() if isinstance(v, datetime.datetime) else v
    s = to_latin(str(v).strip())
    m = DATE_RE.match(s)
    if not m:
        return None
    a, b, y = int(m.group(1)), int(m.group(3)), int(m.group(4))
    if b > 12 and a <= 12:
        month, day = a, b
    else:
        day, month = a, b
    if month < 1 or month > 12 or day < 1 or day > 31:
        return None
    try:
        d = datetime.date(y, month, day)
    except ValueError:
        return None
    if d.month != month or d.day != day:
        return None
    return d

def fmt(d):
    return "%02d/%02d/%04d" % (d.day, d.month, d.year)

# keyword, name_col, permit_col, mailday_col
SHEETS = [
    ("أفراد",  3, 9, 10),
    ("مركبات", 3, 6, 7),
]

wb = openpyxl.load_workbook(SRC)
report = {}

for keyword, ncol, pcol, dcol in SHEETS:
    ws = None
    for sh in wb.worksheets:
        if keyword in sh.title:
            ws = sh; break
    if ws is None:
        report[keyword] = "SHEET NOT FOUND"; continue

    maxc = ws.max_column
    st = {"green_valid": 0, "yellow_expired": 0, "red_ban_rows": 0,
          "grey_awaiting": 0, "white_other": 0,
          "permit_dates_fixed": 0, "mailday_dates_fixed": 0, "separators": 0}
    other = {}

    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=ncol).value
        if name is None or str(name).strip() == "":
            st["separators"] += 1
            continue

        # ---- normalise Mail Day, force white ----
        dcell = ws.cell(row=r, column=dcol)
        md = parse_date(dcell.value)
        if md:
            dcell.value = fmt(md)
            dcell.number_format = "@"
            st["mailday_dates_fixed"] += 1
        dcell.fill = WHITE

        # ---- permit ----
        pcell = ws.cell(row=r, column=pcol)
        praw = pcell.value
        ptext = "" if praw is None else to_latin(str(praw).strip())

        if "منع" in ptext:
            for c in range(1, maxc + 1):
                ws.cell(row=r, column=c).fill = RED
            st["red_ban_rows"] += 1
            continue  # whole row already red; date/other columns stay red

        pd = parse_date(praw)
        if pd:
            pcell.value = fmt(pd)
            pcell.number_format = "@"
            st["permit_dates_fixed"] += 1
            if pd < TODAY:
                pcell.fill = YELLOW; st["yellow_expired"] += 1
            else:
                pcell.fill = GREEN; st["green_valid"] += 1
        elif "انتظار" in ptext:
            pcell.fill = GREY; st["grey_awaiting"] += 1
        else:
            pcell.fill = WHITE; st["white_other"] += 1
            if ptext:
                other[ptext[:24]] = other.get(ptext[:24], 0) + 1

    report[ws.title] = {"stats": st, "white_other_values": other}

wb.save(OUT)
print("TODAY =", TODAY)
print(json.dumps(report, ensure_ascii=False, indent=1))
print("SAVED:", OUT)
