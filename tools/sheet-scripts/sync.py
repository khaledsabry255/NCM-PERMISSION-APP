# -*- coding: utf-8 -*-
"""
Push the desktop workbook to the Google Sheet the site reads.

The sheet is a mirror of C:\\Users\\Security\\Desktop\\work sheet.xlsx - same
values, same colours. This script refuses to upload anything it cannot vouch
for, and never records a push as done unless the sheet actually took it.
"""
import base64, csv, datetime, hashlib, io, json, os, re, shutil, subprocess, sys, urllib.parse

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = r"C:/Users/Security/Desktop/work sheet.xlsx"
WORK = os.path.join(HERE, "upload.xlsx")
STATE = os.path.join(HERE, "state.json")
BACKUP = os.path.join(HERE, "backup")
DESK_BACKUP = r"C:/Users/Security/Desktop/نسخ احتياطية - work sheet"
KEEP_BACKUPS = 10          # each copy is ~20 MB, so the folder is capped
SHEET_ID = "1ZceJtmQMpW7Ky3Ysgz0mDcmE6Uxr600OuO2KOMHJnWE"
# Only these two are read by the site, so only these two are worth stopping an
# upload over. Every other tab rides along exactly as it is - a helper tab being
# renamed, added or merged away is the author's business, not a failure.
REQUIRED = ["\u062a\u0635\u0627\u0631\u064a\u062d \u0623\u0641\u0631\u0627\u062f ",
            "\u062a\u0635\u0627\u0631\u064a\u062d \u0645\u0631\u0643\u0628\u0627\u062a"]
# Tabs the site never reads. They are working areas: a paste buffer and the
# lookup table the ASE columns are filled from. The upload carries the values
# those formulas already produced, so the sheet needs neither - and leaving the
# lookup table out keeps a 247k-row table from being copied on every sync.
# Tried uploading ASE Data alongside it once (2026-09-03) - Google's own
# xlsx-to-Sheets conversion failed on the larger file ("unable to load the
# document"), so it stays dropped.
DROP = ["DB Paste", "DB Key", "ASE Data", "ASE Letters"]
IND, VEH = REQUIRED[0], REQUIRED[1]

# Column order the site reads by position. If this moves the app breaks with no
# error at all, so a mismatch stops the upload.
EXPECT = {
    IND: ["\u0627\u0644\u0627\u0633\u0640\u0640\u0640\u0640\u0640\u0640\u0640\u0640\u0640\u0645",
          "\u0627\u0644\u0648\u0638\u064a\u0641\u0629",
          "\u0627\u0644\u0631\u0642\u0645 \u0627\u0644\u0642\u0648\u0645\u0649",
          "\u0627\u0644\u0639\u0646\u0648\u0627\u0646", "\u0645\u0644\u0627\u062d\u0638\u0627\u062a",
          "\u0627\u0644\u0645\u0631\u0633\u0644", "\u062a\u0635\u0631\u064a\u062d",
          "Mail Day", "MAIL TYPE", "Letter Number"],
    VEH: ["\u0631\u0642\u0645 \u0627\u0644\u0633\u064a\u0627\u0631\u0629", "\u0627\u0644\u0646\u0648\u0639",
          "\u062a\u0627\u0628\u0639\u0647", "\u062a\u0635\u0631\u064a\u062d", "Mail Day", "MAIL TYPE"],
}
GOV = set("01 02 03 04 11 12 13 14 15 16 17 18 19 21 22 23 24 25 26 27 28 29 31 32 33 34 35 88".split())
DATE_RE = re.compile(r"^\s*(\d{1,2})\s*[-/]\s*(\d{1,2})\s*[-/]\s*(\d{4})")


def norm(v):
    return re.sub(r"\s+", " ", str(v if v is not None else "")).strip()


def digits(v):
    return re.sub(r"\D", "", norm(v))


def parse_date(v):
    """Day first unless that is impossible - matches the site's own parser."""
    m = DATE_RE.match(norm(v))
    if not m:
        return None
    a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    mo, d = (a, b) if (b > 12 and a <= 12) else (b, a)
    try:
        return datetime.date(y, mo, d)
    except ValueError:
        return None


def sha256(path, chunk=1 << 20):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for b in iter(lambda: f.read(chunk), b""):
            h.update(b)
    return h.hexdigest()


def gviz(tab):
    url = ("https://docs.google.com/spreadsheets/d/%s/gviz/tq?tqx=out:csv&sheet=%s&cb=%d"
           % (SHEET_ID, urllib.parse.quote(tab), datetime.datetime.now().microsecond))
    out = subprocess.run(["curl", "-sL", "--max-time", "120", url], capture_output=True)
    return list(csv.reader(io.StringIO(out.stdout.decode("utf-8", "replace"))))


def audit(rows, key, permit, day, idcol):
    """rows[0] is the header. Column numbers are 0-based. -> (data_rows, problems)"""
    n = 0
    bad = []
    for i, r in enumerate(rows[1:], start=2):
        g = lambda c: norm(r[c]) if len(r) > c else ""
        if not g(key):
            continue
        n += 1
        v = g(day)
        if v and not parse_date(v):
            bad.append("\u0635\u0641 %d: \u062a\u0627\u0631\u064a\u062e \u0625\u0631\u0633\u0627\u0644 %s" % (i, v))
        m = DATE_RE.match(g(permit))
        if m and not (2020 <= int(m.group(3)) <= 2035):
            bad.append("\u0635\u0641 %d: \u0633\u0646\u0629 \u062a\u0635\u0631\u064a\u062d %s" % (i, g(permit)))
        if idcol is not None:
            d = digits(g(idcol))
            ok = len(d) == 14 and d[0] in "23" and d[7:9] in GOV
            if ok:
                try:
                    datetime.date((1900 if d[0] == "2" else 2000) + int(d[1:3]),
                                  int(d[3:5]), int(d[5:7]))
                except ValueError:
                    ok = False
            if not ok:
                bad.append("\u0635\u0641 %d: \u0631\u0642\u0645 \u0642\u0648\u0645\u0649" % i)
    return n, bad


def die(msg, **extra):
    print(json.dumps(dict(ok=False, error=msg, **extra), ensure_ascii=False))
    sys.exit(1)


def main():
    if not os.path.exists(SRC):
        die("\u0645\u0644\u0641 \u0627\u0644\u062f\u064a\u0633\u0643 \u062a\u0648\u0628 \u0645\u0634 \u0645\u0648\u062c\u0648\u062f: " + SRC)
    for f in ("webapp_url.txt", "secret.txt"):
        if not os.path.exists(os.path.join(HERE, f)):
            die("\u0646\u0627\u0642\u0635 " + f)
    url = open(os.path.join(HERE, "webapp_url.txt"), encoding="utf-8").read().strip()
    secret = open(os.path.join(HERE, "secret.txt"), encoding="utf-8").read().strip()

    digest = sha256(SRC)
    prev = {}
    if os.path.exists(STATE):
        try:
            prev = json.load(open(STATE, encoding="utf-8"))
        except Exception:
            pass
    if prev.get("pushed_sha256") == digest and "--force" not in sys.argv:
        print(json.dumps({"ok": True, "skipped": "no_change"}, ensure_ascii=False))
        return

    # Keep a dated copy of the workbook itself before it goes anywhere, and
    # keep the folder from growing without bound.
    os.makedirs(DESK_BACKUP, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M")
    shutil.copy2(SRC, os.path.join(DESK_BACKUP, "work sheet %s.xlsx" % stamp))
    kept = sorted(f for f in os.listdir(DESK_BACKUP) if f.lower().endswith(".xlsx"))
    for old in kept[:-KEEP_BACKUPS]:
        try: os.remove(os.path.join(DESK_BACKUP, old))
        except OSError: pass

    import openpyxl
    shutil.copy2(SRC, WORK)
    # data_only reads the values Excel already worked out and drops the formulas
    # behind them. The ASE columns are array formulas looking up a 247k-row table
    # in another tab; sending the results means the sheet does not need that
    # table at all, and does not carry formulas that would break without it.
    wb = openpyxl.load_workbook(WORK, data_only=True)
    missing = [t for t in REQUIRED if t not in wb.sheetnames]
    if missing:
        die("tabs_missing", tabs=missing)
    for t in DROP:
        if t in wb.sheetnames:
            del wb[t]

    carried = list(wb.sheetnames)

    # "الاســــــــــم" is padded with tatweel purely for looks and the run
    # length drifts as the header is retyped; it says nothing about the order.
    plain = lambda s: s.replace("ـ", "").strip()
    for tab, expect in EXPECT.items():
        hdr = [norm(c.value) for c in next(wb[tab].iter_rows(min_row=1, max_row=1))]
        got = [plain(h) for h in hdr[2:2 + len(expect)]]
        if got != [plain(e) for e in expect]:
            die("column_order_changed", tab=tab, found=got, expected=[plain(e) for e in expect])

    # A national id pasted as text sits in a numeric column, and gviz then hands
    # the app an empty cell - the person silently stops being searchable.
    ws = wb[IND]
    coerced = 0
    for r in range(2, ws.max_row + 1):
        if not norm(ws.cell(r, 3).value):
            continue
        c = ws.cell(r, 5)
        if isinstance(c.value, str):
            d = digits(c.value)
            if len(d) == 14:
                c.value = int(d)
                c.number_format = "0"
                coerced += 1

    # Every border in the workbook takes its colour from the theme, and Google
    # drops a theme-coloured border on import - the grid Excel draws arrives on
    # Drive with nothing to see. Restating the same colour as plain black keeps
    # it. Only the upload copy is touched; the desktop file keeps its own theme.
    from openpyxl.styles import Border, Side
    from openpyxl.formatting.rule import Rule
    from openpyxl.formatting.formatting import ConditionalFormattingList

    # Styled-but-empty rows run to the sheet's limit. Only the rows that carry a
    # name or a plate are worth walking - Google trims the rest.
    def last_row(sh):
        return max((r for r in range(1, sh.max_row + 1)
                    if norm(sh.cell(r, 3).value)), default=1)

    def blacken(sh, ncol):
        last = last_row(sh)
        for row in sh.iter_rows(min_row=1, max_row=last, max_col=ncol):
            for c in row:
                b = c.border
                if not (b.left.style or b.right.style or b.top.style or b.bottom.style):
                    continue
                black = lambda s: Side(style=s.style, color="FF000000") if s.style else s
                c.border = Border(left=black(b.left), right=black(b.right),
                                  top=black(b.top), bottom=black(b.bottom))
    blacken(wb[IND], 11)
    blacken(wb[VEH], 8)

    # Excel's "duplicate values" rule is the one piece of formatting Google has
    # no import for, so the pink on a repeated national id never reached Drive.
    # The same test written as a formula does convert.
    #
    # Google ignores the priority attribute and ranks rules by the order they
    # appear in the file, first match winning - so the rule is put at the head
    # of the list, not merely given priority 1. That is where Excel ranks it
    # too: a repeated id has to stay visible whatever the permit column says.
    def dup_ids(sh):
        dxf = next((r.dxf for rng in sh.conditional_formatting for r in rng.rules
                    if r.type == "duplicateValues"), None)
        if dxf is None:
            return False
        last = last_row(sh)
        existing = [(str(cf.sqref), list(rules))
                    for cf, rules in sh.conditional_formatting._cf_rules.items()]
        sh.conditional_formatting = ConditionalFormattingList()
        sh.conditional_formatting.add(
            "E2:E%d" % last,
            Rule(type="expression", dxf=dxf, priority=1,
                 formula=["COUNTIF($E$2:$E$%d,$E2)>1" % last]))
        for sqref, rules in existing:
            for r in rules:
                sh.conditional_formatting.add(sqref, r)
        return True

    dup_ids(wb[IND])

    wb.save(WORK)

    wb = openpyxl.load_workbook(WORK, read_only=True, data_only=True)
    counts = {}
    for tab, key, permit, day, idcol in ((IND, 3, 9, 10, 5), (VEH, 3, 6, 7, None)):
        rows = [[norm(v) for v in row] for row in wb[tab].iter_rows(values_only=True)]
        n, bad = audit(rows, key - 1, permit - 1, day - 1,
                       None if idcol is None else idcol - 1)
        if bad:
            die("file_has_problems", tab=tab, count=len(bad), problems=bad[:20])
        counts[tab] = n
    wb.close()

    # Keep what is on Drive before replacing it.
    os.makedirs(BACKUP, exist_ok=True)
    before = {}
    for tab in (IND, VEH):
        rows = gviz(tab)
        before[tab] = sum(1 for r in rows[1:] if len(r) > 2 and norm(r[2]))
        with io.open(os.path.join(BACKUP, tab.strip() + ".csv"), "w",
                     encoding="utf-8", newline="") as f:
            csv.writer(f).writerows(rows)

    # A file that lost most of its rows is a mistake, not an edit.
    for tab in (IND, VEH):
        if before[tab] and counts[tab] < before[tab] * 0.9:
            die("file_much_smaller_than_sheet", tab=tab,
                file_rows=counts[tab], sheet_rows=before[tab])

    b64 = os.path.join(HERE, "payload.b64")
    with open(WORK, "rb") as f, open(b64, "wb") as o:
        o.write(base64.b64encode(f.read()))
    try:
        # No -X POST: it would force POST on the redirect too, with no body and
        # so no Content-Length, which Google answers with 411. 'Expect:' stops
        # curl asking for 100-continue, which Google also answers with 411.
        res = subprocess.run(
            ["curl", "-sL", "--max-time", "600", "%s?secret=%s" % (url, secret),
             "-H", "Expect:", "-H", "Content-Type: text/plain",
             "--data-binary", "@" + b64], capture_output=True)
        body = res.stdout.decode("utf-8", "replace").strip()
    finally:
        if os.path.exists(b64):
            os.remove(b64)

    try:
        reply = json.loads(body)
    except Exception:
        die("bad_reply_from_google", reply=body[:300])
    if not reply.get("ok"):
        die("google_rejected", reply=reply)

    # Read it back: the upload is only done when the sheet really shows it.
    after = {}
    for tab, key, permit, day, idcol in ((IND, 2, 8, 9, 4), (VEH, 2, 5, 6, None)):
        rows = gviz(tab)
        n, bad = audit(rows, key, permit, day, idcol)
        if bad:
            die("sheet_has_problems_after_upload", tab=tab, count=len(bad), problems=bad[:20])
        after[tab] = n

    json.dump({"pushed_sha256": digest, "at": datetime.datetime.now().isoformat()},
              open(STATE, "w", encoding="utf-8"))
    print(json.dumps({
        "ok": True,
        "uploaded": True,
        "file_rows": counts,
        "sheet_rows_before": before,
        "sheet_rows_after": after,
        "ids_coerced": coerced,
        "tabs": carried,
    }, ensure_ascii=False))


main()
